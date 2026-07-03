"""
Add a 'Utilization Effect' tab to Token_and_Data_Build_Out_v4_2.xlsx — a VIEW of the core
engine that makes the ITEM 12 utilization MATURATION visible: utilization climbing from
~10% today toward the agent/human ceiling, and the power it absorbs.

It references the Efficiency Overlay engine for demand (col F), supply (col G) and the
util CEILING (col I), then applies its OWN maturation curve (util_start + maturity-year
cells on this tab, smoothstep) so it ALWAYS shows the effect regardless of the Overlay
scenario picker. Two charts:
  • POWER: power w/o util (= demand) vs power w/ util (after the maturation haircut) vs
    supply. The gap between the two power lines = GW absorbed.
  • UTILIZATION: util ceiling vs util ACTUAL climbing from util_start toward the ceiling.

power w/ util = demand × util_start / util_actual(year);  util_actual re-anchors to
util_start at 2025 (haircut = 1.0 there) and deepens as utilization matures.

  python research/build_utilization_effect_tab.py

SAFETY: surgical zip add of new parts (sheet7 / drawing4 / chart4 / chart5) + edits to
workbook.xml, its rels and [Content_Types]; calcChain dropped + fullCalcOnLoad set; golden
hash re-asserted; never openpyxl-saves. Idempotent. Base + Macro Levers + Overlay untouched.
"""
from __future__ import annotations

import hashlib
import os
import re
import sys
import xml.dom.minidom as minidom
import zipfile
from pathlib import Path

_REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_REPO))
import model  # noqa: E402
import research.build_efficiency_overlay_core as core  # noqa: E402

XLSX = _REPO / "Token_and_Data_Build_Out_v4_2.xlsx"
GOLDEN = "5aba31680bd17859"
SHEET = "Utilization Effect"
OV = "Efficiency Overlay"
YEARS = core.YEARS
N = len(YEARS)
NR0 = 10                     # first data row on this tab
USTART = 0.10               # default 2025 actual utilization (editable on the tab, B4)
UMAT = 2030                 # default maturity year (editable on the tab, D4)

# style indices resolved from styles.xml at runtime (placeholders)
S_TITLE = S_ITAL = S_LBL = S_GEN = S_COLHDR = S_BANNER = 0
S_GW = S_PCT = S_YR = S_KGW = S_KYR = S_KPCT = S_INP_PCT = S_INP_YR = 0

esc = lambda s: str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _resolve_styles():
    z = zipfile.ZipFile(XLSX)
    sm = core.style_map(z.read("xl/styles.xml").decode())
    z.close()
    globals().update(
        S_TITLE=sm["title"], S_ITAL=sm["ital"], S_LBL=sm["lbl"], S_GEN=sm["gen"],
        S_COLHDR=sm["colhdr"], S_BANNER=sm["banner"], S_GW=sm["gw"], S_PCT=sm["pct"],
        S_YR=sm["yr"], S_KGW=sm["gw_bold"], S_KYR=sm["yr_bold"], S_KPCT=sm["pct_bold"],
        S_INP_PCT=sm["pct_in"], S_INP_YR=sm["picker"])


def fmt(v):
    if isinstance(v, int):
        return str(v)
    return repr(float(v))


def _s(s):
    return f' s="{s}"' if s is not None else ""


def cs(ref, text, s=None):
    return f'<c r="{ref}"{_s(s)} t="inlineStr"><is><t>{esc(text)}</t></is></c>'


def cn(ref, val, s=None):
    return f'<c r="{ref}"{_s(s)}><v>{fmt(val)}</v></c>'


def cf(ref, formula, cache, s=None):
    f = formula[1:] if formula.startswith("=") else formula
    if isinstance(cache, str):
        return f'<c r="{ref}"{_s(s)} t="str"><f>{esc(f)}</f><v>{esc(cache)}</v></c>'
    return f'<c r="{ref}"{_s(s)}><f>{esc(f)}</f><v>{fmt(cache)}</v></c>'


def _maturity(y):
    if UMAT <= 2025:
        return 1.0
    t = min(1.0, max(0.0, (y - 2025) / (UMAT - 2025)))
    return t * t * (3.0 - 2.0 * t)


def compute():
    """Cached values: applies ITEM 12 maturation (util_start -> ceiling) to the Base
    demand/ceiling from the core reference, so the tab shows the effect by default."""
    R = core.reference()
    fl, su, ceil = R["floored"], R["supply"], R["ceiling"]
    actual = {y: USTART + (ceil[y] - USTART) * _maturity(y) for y in YEARS}
    pwr_off = {y: fl[y] for y in YEARS}                          # demand, no util credit
    pwr_on = {y: fl[y] * actual[2025] / actual[y] for y in YEARS}  # after maturation haircut
    absorbed = {y: pwr_off[y] - pwr_on[y] for y in YEARS}
    abs_pct = {y: (absorbed[y] / pwr_off[y] if pwr_off[y] else 0.0) for y in YEARS}
    peak_abs = max(absorbed.values())
    peak_abs_y = [y for y in YEARS if absorbed[y] == peak_abs][0]
    return dict(off=pwr_off, on=pwr_on, sup=su, ceil=ceil, act=actual, ab=absorbed, abp=abs_pct,
                peak_abs=peak_abs, peak_abs_y=peak_abs_y, cum=sum(absorbed.values()))


def build_sheet_xml(C):
    rows = {}
    last = NR0 + N - 1
    rows[1] = [cs("A1", "UTILIZATION EFFECT  —  utilization maturing over time, and the power it absorbs", S_TITLE)]
    rows[2] = [cs("A2", "A VIEW of the core engine. 'Power w/o util' = Efficiency Overlay demand (no utilization credit). 'Power w/ util' = the same "
                        "demand after the ITEM 12 maturation haircut: utilization climbs from util_start (B4) toward the agent/human CEILING by the "
                        "maturity year (D4). The gap between the two power lines = GW absorbed by running the fleet hotter. Edit B4 / D4 to reshape the "
                        "climb. (Demand/ceiling track Efficiency Overlay!B4; the maturation is applied here so the effect always shows.)", S_ITAL)]

    rows[4] = [cs("A4", "util_start — 2025 ACTUAL utilization", S_LBL), cn("B4", USTART, S_INP_PCT),
               cs("C4", "matures to ceiling by (year) ▶", S_LBL), cn("D4", UMAT, S_INP_YR)]

    rows[6] = [cs("A6", "ABSORPTION (live)", S_BANNER), cs("B6", "", S_BANNER), cs("C6", "", S_BANNER), cs("D6", "", S_BANNER)]
    rows[7] = [cs("A7", "Peak absorption (GW)", S_LBL), cf("B7", f"=MAX(G{NR0}:G{last})", C["peak_abs"], S_KGW),
               cs("C7", "in year ▶", S_LBL),
               cf("D7", f"=INDEX(A{NR0}:A{last},MATCH(MAX(G{NR0}:G{last}),G{NR0}:G{last},0))", C["peak_abs_y"], S_KYR)]
    rows[8] = [cs("A8", "2042 absorption (GW)", S_LBL), cf("B8", f"=G{last}", C["ab"][2042], S_KGW),
               cs("C8", "of 2042 demand ▶", S_LBL), cf("D8", f"=H{last}", C["abp"][2042], S_KPCT)]

    hdr = NR0 - 1
    rows[hdr] = [cs(f"A{hdr}", "year", S_COLHDR), cs(f"B{hdr}", "power w/o util GW", S_COLHDR),
                 cs(f"C{hdr}", "util ceiling", S_COLHDR), cs(f"D{hdr}", "util ACTUAL", S_COLHDR),
                 cs(f"E{hdr}", "power w/ util GW", S_COLHDR), cs(f"F{hdr}", "supply GW", S_COLHDR),
                 cs(f"G{hdr}", "absorbed GW", S_COLHDR), cs(f"H{hdr}", "absorbed %", S_COLHDR)]

    for idx, y in enumerate(YEARS):
        r = NR0 + idx
        oc = core.yc(y)                          # Efficiency Overlay column for this year (transposed)
        q = f"'{OV}'!"
        tx = f"MIN(1,MAX(0,(A{r}-2025)/($D$4-2025)))"
        rows[r] = [
            cf(f"A{r}", f"={q}{oc}{core.YHDR}", int(y), S_YR),          # year (flow header row)
            cf(f"B{r}", f"={q}{oc}{core.R_DEM}", C["off"][y], S_GW),    # demand (flow row)
            cf(f"C{r}", f"={q}{oc}{core.R_CEIL}", C["ceil"][y], S_PCT), # util ceiling (flow row)
            cf(f"D{r}", f"=$B$4+(C{r}-$B$4)*({tx})^2*(3-2*({tx}))", C["act"][y], S_PCT),
            cf(f"E{r}", f"=B{r}*$B$4/D{r}", C["on"][y], S_GW),
            cf(f"F{r}", f"={q}{oc}{core.R_SUP}", C["sup"][y], S_GW),    # supply (flow row)
            cf(f"G{r}", f"=B{r}-E{r}", C["ab"][y], S_GW),
            cf(f"H{r}", f"=IF(B{r}=0,0,G{r}/B{r})", C["abp"][y], S_PCT),
        ]

    body = "".join(f'<row r="{n}">{"".join(rows[n])}</row>' for n in sorted(rows))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<dimension ref="A1:H{last}"/><sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/>'
        '<cols><col min="1" max="1" width="32"/><col min="2" max="8" width="15"/></cols>'
        '<sheetData>' + body + '</sheetData>'
        '<drawing r:id="rId1"/></worksheet>'
    )


def num_cache(vals):
    pts = "".join(f'<c:pt idx="{i}"><c:v>{fmt(v)}</c:v></c:pt>' for i, v in enumerate(vals))
    return f'<c:numCache><c:formatCode>General</c:formatCode><c:ptCount val="{len(vals)}"/>{pts}</c:numCache>'


def series(idx, title_ref, title_cache, col, vals, cats_vals, color):
    last = NR0 + N - 1
    cat_pts = "".join(f'<c:pt idx="{i}"><c:v>{int(v)}</c:v></c:pt>' for i, v in enumerate(cats_vals))
    return (
        f'<c:ser><c:idx val="{idx}"/><c:order val="{idx}"/>'
        f'<c:tx><c:strRef><c:f>{title_ref}</c:f><c:strCache><c:ptCount val="1"/><c:pt idx="0"><c:v>{esc(title_cache)}</c:v></c:pt></c:strCache></c:strRef></c:tx>'
        f'<c:spPr><a:ln w="28575"><a:solidFill><a:srgbClr val="{color}"/></a:solidFill></a:ln></c:spPr>'
        f'<c:marker><c:symbol val="none"/></c:marker>'
        f"<c:cat><c:numRef><c:f>'{SHEET}'!$A${NR0}:$A${last}</c:f>"
        f'<c:numCache><c:formatCode>General</c:formatCode><c:ptCount val="{N}"/>{cat_pts}</c:numCache></c:numRef></c:cat>'
        f"<c:val><c:numRef><c:f>'{SHEET}'!${col}${NR0}:${col}${last}</c:f>{num_cache(vals)}</c:numRef></c:val>"
        f'<c:smooth val="0"/></c:ser>'
    )


def _chart(title, sers, cx, vx, ymax=None, ypct=False):
    ax_fmt = '<c:numFmt formatCode="0%" sourceLinked="0"/>' if ypct else ''
    scaling = '<c:scaling><c:orientation val="minMax"/>' + (f'<c:max val="{ymax}"/>' if ymax else '') + '<c:min val="0"/></c:scaling>'
    ytitle = 'utilization' if ypct else 'GW'
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        '<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<c:chart>'
        f'<c:title><c:tx><c:rich><a:bodyPr/><a:p><a:r><a:t>{esc(title)}</a:t></a:r></a:p></c:rich></c:tx><c:overlay val="0"/></c:title>'
        '<c:autoTitleDeleted val="0"/><c:plotArea><c:layout/>'
        '<c:lineChart><c:grouping val="standard"/><c:varyColors val="0"/>'
        + "".join(sers) +
        '<c:marker val="1"/>'
        f'<c:axId val="{cx}"/><c:axId val="{vx}"/></c:lineChart>'
        f'<c:catAx><c:axId val="{cx}"/><c:scaling><c:orientation val="minMax"/></c:scaling><c:delete val="0"/>'
        f'<c:axPos val="b"/><c:title><c:tx><c:rich><a:bodyPr/><a:p><a:r><a:t>year</a:t></a:r></a:p></c:rich></c:tx><c:overlay val="0"/></c:title>'
        f'<c:crossAx val="{vx}"/></c:catAx>'
        f'<c:valAx><c:axId val="{vx}"/>{scaling}<c:delete val="0"/><c:axPos val="l"/>{ax_fmt}'
        f'<c:title><c:tx><c:rich><a:bodyPr rot="-5400000" vert="horz"/><a:p><a:r><a:t>{ytitle}</a:t></a:r></a:p></c:rich></c:tx><c:overlay val="0"/></c:title>'
        f'<c:crossAx val="{cx}"/></c:valAx></c:plotArea>'
        '<c:legend><c:legendPos val="b"/><c:overlay val="0"/></c:legend>'
        '<c:plotVisOnly val="1"/><c:dispBlanksAs val="gap"/></c:chart></c:chartSpace>'
    )


def build_charts(C):
    yrs = list(YEARS)
    # power chart (chart4): demand vs realized vs supply; cap axis (supply runs away to ~1858)
    pw = [series(0, f"'{SHEET}'!$B$9", "power w/o util GW", "B", [C["off"][y] for y in yrs], yrs, "C00000"),
          series(1, f"'{SHEET}'!$E$9", "power w/ util GW", "E", [C["on"][y] for y in yrs], yrs, "2E7D32"),
          series(2, f"'{SHEET}'!$F$9", "supply GW", "F", [C["sup"][y] for y in yrs], yrs, "808080")]
    power = _chart("Power demand: with vs without utilization (gap = absorbed)", pw, 111111111, 222222222, ymax=1000)
    # util chart (chart5): ceiling vs actual climbing
    uw = [series(0, f"'{SHEET}'!$C$9", "util ceiling", "C", [C["ceil"][y] for y in yrs], yrs, "1565C0"),
          series(1, f"'{SHEET}'!$D$9", "util ACTUAL", "D", [C["act"][y] for y in yrs], yrs, "EF6C00")]
    util = _chart("Utilization maturing: actual climbs from util_start toward the ceiling", uw, 333333333, 444444444, ymax=0.8, ypct=True)
    return power, util


DRAWING = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
    'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
    'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
    # power chart
    '<xdr:twoCellAnchor editAs="oneCell">'
    '<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>28</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
    '<xdr:to><xdr:col>8</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>48</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
    '<xdr:graphicFrame macro=""><xdr:nvGraphicFramePr><xdr:cNvPr id="2" name="Power chart"/>'
    '<xdr:cNvGraphicFramePr/></xdr:nvGraphicFramePr><xdr:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></xdr:xfrm>'
    '<a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/chart">'
    '<c:chart xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" '
    'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1"/>'
    '</a:graphicData></a:graphic></xdr:graphicFrame><xdr:clientData/></xdr:twoCellAnchor>'
    # util chart
    '<xdr:twoCellAnchor editAs="oneCell">'
    '<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>49</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
    '<xdr:to><xdr:col>8</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>67</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
    '<xdr:graphicFrame macro=""><xdr:nvGraphicFramePr><xdr:cNvPr id="3" name="Utilization chart"/>'
    '<xdr:cNvGraphicFramePr/></xdr:nvGraphicFramePr><xdr:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></xdr:xfrm>'
    '<a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/chart">'
    '<c:chart xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" '
    'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId2"/>'
    '</a:graphicData></a:graphic></xdr:graphicFrame><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>'
)

DRAWING_RELS = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" Target="../charts/chart4.xml"/>'
    '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" Target="../charts/chart5.xml"/>'
    '</Relationships>'
)

SHEET_RELS = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing4.xml"/>'
    '</Relationships>'
)


def main():
    _resolve_styles()
    C = compute()
    sheet_xml = build_sheet_xml(C)
    chart_power, chart_util = build_charts(C)
    for blob in (sheet_xml, chart_power, chart_util, DRAWING, DRAWING_RELS, SHEET_RELS):
        minidom.parseString(blob)

    z = zipfile.ZipFile(XLSX)
    wb = z.read("xl/workbook.xml").decode()
    rels = z.read("xl/_rels/workbook.xml.rels").decode()
    ct = z.read("[Content_Types].xml").decode()

    wb = re.sub(r'<sheet name="%s"[^>]*/>' % re.escape(SHEET), "", wb)
    new_sheet = f'<sheet name="{SHEET}" sheetId="7" r:id="rId12"/>'
    wb = re.sub(r'(<sheet name="%s"[^>]*/>)' % re.escape(OV), r"\1" + new_sheet, wb, count=1)
    if "fullCalcOnLoad" not in wb:
        wb = re.sub(r'<calcPr ([^>]*?)/>', r'<calcPr \1 fullCalcOnLoad="1"/>', wb)

    rels = re.sub(r'<Relationship Id="rId12"[^>]*/>', "", rels)
    rels = rels.replace(
        "</Relationships>",
        '<Relationship Id="rId12" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet7.xml"/></Relationships>')
    rels = re.sub(r'<Relationship [^>]*Target="calcChain\.xml"[^>]*/>', "", rels)

    for part, ctype in [
        ("/xl/worksheets/sheet7.xml", "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"),
        ("/xl/drawings/drawing4.xml", "application/vnd.openxmlformats-officedocument.drawing+xml"),
        ("/xl/charts/chart4.xml", "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"),
        ("/xl/charts/chart5.xml", "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"),
    ]:
        ct = re.sub(r'<Override PartName="%s"[^>]*/>' % re.escape(part), "", ct)
        ct = ct.replace("</Types>", f'<Override PartName="{part}" ContentType="{ctype}"/></Types>')
    ct = re.sub(r'<Override PartName="/xl/calcChain\.xml"[^>]*/>', "", ct)

    for blob in (wb, rels, ct):
        minidom.parseString(blob)

    new_parts = {
        "xl/worksheets/sheet7.xml": sheet_xml,
        "xl/worksheets/_rels/sheet7.xml.rels": SHEET_RELS,
        "xl/drawings/drawing4.xml": DRAWING,
        "xl/drawings/_rels/drawing4.xml.rels": DRAWING_RELS,
        "xl/charts/chart4.xml": chart_power,
        "xl/charts/chart5.xml": chart_util,
    }
    edited = {"xl/workbook.xml": wb, "xl/_rels/workbook.xml.rels": rels, "[Content_Types].xml": ct}
    drop = {"xl/calcChain.xml"}

    tmp = str(XLSX) + ".NEW.xlsx"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for item in z.infolist():
            if item.filename in drop or item.filename in new_parts:
                continue
            if item.filename in edited:
                out.writestr(item, edited[item.filename].encode())
            else:
                out.writestr(item, z.read(item.filename))
        for name, blob in new_parts.items():
            out.writestr(name, blob.encode())
    z.close()

    assert zipfile.ZipFile(tmp).testzip() is None
    import openpyxl
    chk = openpyxl.load_workbook(tmp, data_only=True)
    assert SHEET in chk.sheetnames and OV in chk.sheetnames and "Base" in chk.sheetnames

    model.EXCEL_PATH = tmp
    mv = model.build_macro_gap(model.build_token_demand("Base"))
    cols = ["gross_tokens_T", "fleet_eff_idx", "net_compute_demand_T", "demand_gw",
            "supply_gw", "gap_gw", "cumulative_power_grid_capex_b"]
    h = hashlib.sha256(mv[cols].round(6).to_csv().encode()).hexdigest()[:16]
    assert h == GOLDEN, f"GOLDEN MISMATCH {h}"
    os.replace(tmp, XLSX)
    model.EXCEL_PATH = XLSX
    print(f"OK: '{SHEET}' tab + 2 charts added; sheets now {chk.sheetnames}")
    print(f"    util_start {USTART:.0%} -> ceiling by {UMAT}; peak absorption {C['peak_abs']:.0f} GW @ {C['peak_abs_y']}; "
          f"2042 absorption {C['ab'][2042]:.0f} GW ({C['abp'][2042]*100:.0f}% of demand); golden {h} intact")


if __name__ == "__main__":
    main()
