"""
Rebuild the 'Efficiency Overlay' sheet as the SINGLE CORE model (token x efficiency
basis), mirroring model.py: tokens -> demand -> supply -> AGENT/HUMAN UTILIZATION ->
realized power -> GAP, with a Bull/Base/Bear/Central scenario picker.

This version fixes three usability problems:
  1. HEADLINE AT THE TOP. The KPI band (peak gap + year, power drawn / supply at the
     peak, 2042 power, balance / overshoot / asymptote years) sits in rows 7-13, right
     under the scenario picker (B4) — no scrolling past the engine to find peak demand.
  2. PER-YEAR UTILIZATION KNOB. The engine has a 'util override' column (J): type a %
     for any year to force the fleet utilization that year; blank = the model's
     agent/human blend. It bites whenever the haircut is on (Central/Bear) OR you've
     typed an override, so you can shape utilization by hand in any scenario.
  3. NUMBER FORMATS. GW as #,##0, util/agent-share as 0.0%, eff as 0.0, years as plain
     integers; bold KPI values, a header fill on the table/column headers, and a yellow
     input fill on the picker + override cells. Two fills + three cellXfs are appended to
     styles.xml (idempotent) — values are untouched, so the golden hash is unaffected.

Utilization (ITEM 11) = demand-weighted blend of a human regime (util_human, cool,
latency-bound) and an agent regime (util_agent, hot, batches). agent_share =
(robotics + enterprise*(1-1/agent_mult)) / total, pulled live from the Base sheet.
realized power = demand * util_used(2025)/util_used(year). The SAME agent multiplier
that drives demand also lifts utilization -> coupled by construction.

Base (picker=2) = model.py committed base (util OFF, no override) -> golden hash tie-out.

  python research/build_efficiency_overlay_core.py

SAFETY: surgical zip overwrite of the sheet + styles.xml; golden hash asserted; Base
column tied out against model.build_macro_gap('Base'); never openpyxl-saves the workbook.
"""
from __future__ import annotations

import hashlib
import os
import re
import sys
import xml.dom.minidom as minidom
import zipfile
from pathlib import Path

from openpyxl.utils import get_column_letter

_REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_REPO))
import model  # noqa: E402

XLSX = _REPO / "Token_and_Data_Build_Out_v4_2.xlsx"
SHEET = "Efficiency Overlay"
GOLDEN = "5aba31680bd17859"
YEARS = list(model.YEARS)
N = len(YEARS)
ANCHOR = 70.0
ALGO = 0.10   # ITEM 2b canonical algorithmic efficiency /yr (owner 2026-07-03); matches model.py default
PHASE = (0.22, 0.30, 0.25, 0.15)
SW_START = 2034
PICKER_DEFAULT = 2
OVERRIDES = {}                    # per-year util overrides preserved from the workbook (year -> %)

# Scenario presets (Bull / Base / Bear / Central). Base == model.py committed defaults.
SCEN = {                          #         Bull    Base    Bear   Central
    "doubling_years":   (3.0,   2.5,   1.40,  2.5),    # PURE hardware doubling (yrs); algo 10%/yr separate (ALGO const)  [owner 2026-07-03: Base/Central 2.0->2.5]
    "fleet_lag_years":  (8,     5,     4,     5),       # fleet refresh lag (yrs)
    "supply_scale":     (0.85,  1.0,   1.15,  1.0),     # x phase rates
    "retire_rate":      (0.0,   0.0,   0.08,  0.0),     # ITEM 9 floor decay /yr
    "enable_split":     (0,     0,     1,     1),        # ITEM 11 agent/human util haircut on/off
    "util_human":       (0.45,  0.45,  0.45,  0.45),    # interactive ceiling
    "util_agent":       (0.75,  0.75,  0.75,  0.75),    # batch/agentic ceiling
    "util_start":       (0.10,  0.10,  0.10,  0.10),    # ITEM 12 2025 ACTUAL util — ALWAYS forecast the climb
    "util_maturity_yr": (2030,  2030,  2030,  2030),    # ITEM 12 year util reaches the ceiling (smoothstep)
    "second_wave":      (0.0,   0.0,   0.0,   0.0),     # ITEM 7 demand re-accel /yr from SW_START
}
SKEYS = list(SCEN.keys())
NK = len(SKEYS)

# ---- row layout (TRANSPOSED engine: line items down rows, years across columns) ----
SI_HDR = 15                       # 'SCENARIO INPUTS' header
SK0 = 16                          # first SCEN-key input row
SW_ROW = SK0 + NK                 # second-wave start year
PR_HDR = SW_ROW + 1               # 'SUPPLY PHASE RATES' header
PR0 = PR_HDR + 1                  # first phase-rate row (4 rows)
ACT_HDR = PR0 + 4 + 1             # 'ACTIVE' header (one blank row before)
ACT0 = ACT_HDR + 1                # first active-pick row
# THE FLOW (waterfall): each step is a ROW, years 2025..2042 march across columns B..S
FLOW_BANNER = ACT0 + NK + 1       # 'THE FLOW' banner (one blank row before)
YHDR = FLOW_BANNER + 1            # year header row (B..S = 2025..2042)
R_DEMHDR = YHDR + 1               # 'DEMAND' sub-label
R_TOK, R_NEWEFF, R_FLEET, R_RAW, R_DEM = YHDR + 2, YHDR + 3, YHDR + 4, YHDR + 5, YHDR + 6
R_SUPHDR = YHDR + 7               # 'SUPPLY' sub-label
R_SUP = YHDR + 8
R_UTILHDR = YHDR + 9              # 'UTILIZATION' sub-label
R_AGENT, R_CEIL, R_ACTUAL, R_OVR, R_USED = YHDR + 10, YHDR + 11, YHDR + 12, YHDR + 13, YHDR + 14
R_RESHDR = YHDR + 15             # 'RESULT' sub-label
R_REAL, R_GAP = YHDR + 16, YHDR + 17
R_BAL, R_OVS = YHDR + 18, YHDR + 19   # helper rows (balance / overshoot flags), hidden-ish


def yc(year):
    """Column letter for a given year in the transposed flow (B=2025 .. S=2042)."""
    return get_column_letter(2 + (year - 2025))


LASTCOL = yc(YEARS[-1])

# Base-sheet row map (for live agent_share + token links).
B_TOTAL, B_ENT, B_RETAIL, B_ROBO, B_MULT = 7, 38, 9, 58, 42

# ---- styles -----------------------------------------------------------------------
# Rather than trust fragile pre-existing cellXfs indices (the workbook's numFmt table has
# been remapped by prior edits), we append a SELF-CONTAINED block of all styles we use,
# with our OWN numFmts (181=0.0%, 182=0.0; builtins 3=#,##0, 1=integer, 0=General) and
# reference fonts (0 normal, 1 bold, 3 bold18, 4 italic) + fills (blue header, yellow input)
# which are stable. ensure_styles() resolves the base index at build time; the S_* globals
# below are placeholders, overwritten in main() once the base is known.
STYLE_ORDER = ["gw", "gw_bold", "pct", "pct_in", "pct_bold", "eff", "yr", "yr_bold",
               "colhdr", "banner", "picker", "title", "ital", "lbl", "gen"]
CUSTOM_N = len(STYLE_ORDER)
S_GW = S_KGW = S_PCT = S_INP_PCT = S_KPCT = S_EFF = S_YR = S_KYR = 0
S_COLHDR = S_BANNER = S_INP_PICK = S_TITLE = S_ITAL = S_LBL = S_GEN = 0


def _custom_xfs(blue, yel):
    return [
        '<xf numFmtId="3" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>',                  # gw  #,##0
        '<xf numFmtId="3" fontId="1" fillId="0" borderId="0" xfId="0" applyNumberFormat="1" applyFont="1"/>',    # gw_bold
        '<xf numFmtId="181" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>',                # pct 0.0%
        f'<xf numFmtId="181" fontId="0" fillId="{yel}" borderId="0" xfId="0" applyNumberFormat="1" applyFill="1"/>',  # pct_in (yellow)
        '<xf numFmtId="181" fontId="1" fillId="0" borderId="0" xfId="0" applyNumberFormat="1" applyFont="1"/>',  # pct_bold
        '<xf numFmtId="182" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>',                # eff 0.0
        '<xf numFmtId="1" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>',                  # yr  integer
        '<xf numFmtId="1" fontId="1" fillId="0" borderId="0" xfId="0" applyNumberFormat="1" applyFont="1"/>',    # yr_bold
        f'<xf numFmtId="0" fontId="1" fillId="{blue}" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="center"/></xf>',  # colhdr
        f'<xf numFmtId="0" fontId="1" fillId="{blue}" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="left"/></xf>',    # banner
        f'<xf numFmtId="1" fontId="1" fillId="{yel}" borderId="0" xfId="0" applyNumberFormat="1" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="center"/></xf>',  # picker
        '<xf numFmtId="0" fontId="3" fillId="0" borderId="0" xfId="0" applyFont="1"/>',                          # title bold18
        '<xf numFmtId="0" fontId="4" fillId="0" borderId="0" xfId="0" applyFont="1"/>',                          # ital
        '<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/>',                          # lbl bold
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>',                                        # gen
    ]


def ensure_styles(s: str):
    """Append our self-contained style block (idempotent). Returns (new_styles, base_index)
    where base_index is the cellXfs index of the first custom xf (STYLE_ORDER[0])."""
    # custom numFmts 181 (0.0%) + 182 (0.0)
    if 'numFmtId="181"' not in s:
        nf = '<numFmt numFmtId="181" formatCode="0.0%"/><numFmt numFmtId="182" formatCode="0.0"/>'
        m = re.search(r'<numFmts count="(\d+)">(.*?)</numFmts>', s, re.S)
        if m:
            s = s[:m.start()] + f'<numFmts count="{int(m.group(1)) + 2}">' + m.group(2) + nf + '</numFmts>' + s[m.end():]
        else:
            s = s.replace('<fonts', f'<numFmts count="2">{nf}</numFmts><fonts', 1)
    # fills: blue header + yellow input
    if 'FFD9E1F2' not in s:
        m = re.search(r'<fills count="(\d+)">(.*?)</fills>', s, re.S)
        nf = ('<fill><patternFill patternType="solid"><fgColor rgb="FFD9E1F2"/></patternFill></fill>'
              '<fill><patternFill patternType="solid"><fgColor rgb="FFFFF2CC"/></patternFill></fill>')
        s = s[:m.start()] + f'<fills count="{int(m.group(1)) + 2}">' + m.group(2) + nf + '</fills>' + s[m.end():]
    fills = re.findall(r'<fill>.*?</fill>', re.search(r'<fills count="\d+">(.*?)</fills>', s, re.S).group(1))
    blue = next(i for i, f in enumerate(fills) if 'FFD9E1F2' in f)
    yel = next(i for i, f in enumerate(fills) if 'FFF2CC' in f)
    # cellXfs: strip any prior custom block (our xfs are the only ones using numFmt 181/182,
    # and they sit contiguously at the end), then append fresh.
    m = re.search(r'<cellXfs count="(\d+)">(.*?)</cellXfs>', s, re.S)
    xfs = re.findall(r'<xf\b[^>]*/>|<xf\b[^>]*>.*?</xf>', m.group(2), re.S)
    if any(('numFmtId="181"' in x or 'numFmtId="182"' in x) for x in xfs):
        xfs = xfs[:len(xfs) - CUSTOM_N]                  # drop our previous block
    base = len(xfs)
    xfs = xfs + _custom_xfs(blue, yel)
    block = f'<cellXfs count="{len(xfs)}">' + "".join(xfs) + '</cellXfs>'
    s = s[:m.start()] + block + s[m.end():]
    return s, base


def style_map(styles_xml: str) -> dict:
    """Resolve {style_name: cellXfs index} from a styles.xml (idempotent). Used by both this
    builder and the Utilization Effect tab builder so they share the same custom style block."""
    _, base = ensure_styles(styles_xml)
    return {n: base + i for i, n in enumerate(STYLE_ORDER)}


def _set_style_globals(base):
    g = globals()
    names = ["gw", "gw_bold", "pct", "pct_in", "pct_bold", "eff", "yr", "yr_bold",
             "colhdr", "banner", "picker", "title", "ital", "lbl", "gen"]
    idx = {n: base + i for i, n in enumerate(names)}
    g.update(S_GW=idx["gw"], S_KGW=idx["gw_bold"], S_PCT=idx["pct"], S_INP_PCT=idx["pct_in"],
             S_KPCT=idx["pct_bold"], S_EFF=idx["eff"], S_YR=idx["yr"], S_KYR=idx["yr_bold"],
             S_COLHDR=idx["colhdr"], S_BANNER=idx["banner"], S_INP_PICK=idx["picker"],
             S_TITLE=idx["title"], S_ITAL=idx["ital"], S_LBL=idx["lbl"], S_GEN=idx["gen"])

CITES = [
    ("[1]", "Efficiency: hardware doubling 2.0yr (owner 2026-06-25; Epoch AI Jun-2025 ~1.85yr rounded up) x algorithmic 10%/yr compounding (owner 2026-07-03; software/model gains, un-lagged)."),
    ("[3]", "Fleet refresh lag ~5yr: hyperscaler GPU refresh 4-5yr (Blackwell->Rubin); cluster lifecycle 5-7yr (JLL / operators)."),
    ("[4-6]", "2025 anchor 70 GW: Cushman & Wakefield 2024-25 (US 40.6 + EMEA 10.3 + APAC 12.2) + JLL/GS mid."),
    ("[7]", "Supply ~200 GW by 2030: JLL 2026 Global DC Outlook (+97 GW 2025-30, ~doubling; 17% Americas CAGR)."),
    ("[8]", "Grid constraint: JLL 2026 (grid-connect waits >4yr); C&W (power delivery 5yr+) caps near-term supply."),
    ("[10]", "AI workload share -> 50% of new builds by 2030; inference overtakes training late 2026 (JLL 2026)."),
    ("[11]", "Utilization ceilings: queueing (response ~ 1/(1-rho)) caps interactive ~40-50%; batch/agentic ~70-80%."),
]

esc = lambda s: str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def fmt(v):
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, int):
        return str(v)
    return repr(float(v))


def _s(s):
    return f' s="{s}"' if s is not None else ""


def cs(ref, text, s=None):
    return f'<c r="{ref}"{_s(s)} t="inlineStr"><is><t>{esc(text)}</t></is></c>'


def cn(ref, val, s=None):
    return f'<c r="{ref}"{_s(s)}><v>{fmt(val)}</v></c>'


def ce(ref, s):
    """Empty styled cell (an editable input)."""
    return f'<c r="{ref}" s="{s}"/>'


def cf(ref, formula, cache, s=None):
    f = formula[1:] if formula.startswith("=") else formula
    if isinstance(cache, str):
        return f'<c r="{ref}"{_s(s)} t="str"><f>{esc(f)}</f><v>{esc(cache)}</v></c>'
    return f'<c r="{ref}"{_s(s)}><f>{esc(f)}</f><v>{fmt(cache)}</v></c>'


def bcol(year):
    return get_column_letter(2 + (year - 2025))   # Base sheet column for a given year


def reference():
    """Compute the Base-scenario engine (picker default) by mirroring the sheet formulas,
    reading token inputs from the Base sheet. demand_gw / gap tie out to model.py."""
    df = model.load_excel_scenario("Base")
    g = lambda label, y: float(df.loc[label, y])
    base_users = g("AI Users (M's)", 2025)
    user_scale = 1400.0 / base_users
    ent = {y: g("Enterprise Used Per Day (T)", y) for y in YEARS}
    rob = {y: g("Usage per day (token) (T's)", y) for y in YEARS}
    retail = {y: g("Retail", y) * user_scale for y in YEARS}
    mult = {y: g("Agent Multiplier", y) for y in YEARS}
    tot = {y: retail[y] + ent[y] + rob[y] for y in YEARS}   # == model.py total_T

    doubling, lag, sscale = SCEN["doubling_years"][1], SCEN["fleet_lag_years"][1], SCEN["supply_scale"][1]
    retire, split = SCEN["retire_rate"][1], SCEN["enable_split"][1]
    uh, ua, swg = SCEN["util_human"][1], SCEN["util_agent"][1], SCEN["second_wave"][1]
    ustart, umat = SCEN["util_start"][1], SCEN["util_maturity_yr"][1]

    new_eff = {y: 2.0 ** ((y - 2025) / doubling) for y in YEARS}
    fleet_eff = {}
    for y in YEARS:
        fleet_eff[y] = 1.0 if y == 2025 else (1 - 1.0 / lag) * fleet_eff[y - 1] + (1.0 / lag) * new_eff[y]
    # ITEM 2b: algorithmic efficiency multiplies the hardware fleet index (deploys fleet-wide
    # immediately, no refresh lag). raw demand = tokens / (fleet_hw x algo_factor), re-anchored.
    raw = {y: (tot[y] / (fleet_eff[y] * (1 + ALGO) ** (y - 2025))) / (tot[2025] / 1.0) * ANCHOR for y in YEARS}

    floored = {}
    for y in YEARS:
        if y == 2025:
            floored[y] = raw[y]
            continue
        terms = [raw[y], floored[y - 1] * (1 - retire)]
        if y >= SW_START:
            gr = (tot[y] - tot[y - 1]) / tot[y - 1]
            terms.append(floored[y - 1] * (1 + min(gr, swg)))
        floored[y] = max(terms)

    def rate(y):
        r = PHASE[0] if y <= 2027 else PHASE[1] if y <= 2030 else PHASE[2] if y <= 2035 else PHASE[3]
        return r * sscale
    supply = {}
    for y in YEARS:
        supply[y] = ANCHOR if y == 2025 else supply[y - 1] * (1 + rate(y))

    a_share = {y: (rob[y] + ent[y] * (1 - 1 / mult[y])) / tot[y] for y in YEARS}
    ceiling = {y: a_share[y] * ua + (1 - a_share[y]) * uh for y in YEARS}    # col I — achievable ceiling

    def maturity(y):                                                         # ITEM 12 smoothstep S-curve
        if umat <= 2025:
            return 1.0
        t = min(1.0, max(0.0, (y - 2025) / (umat - 2025)))
        return t * t * (3.0 - 2.0 * t)
    # col J — ACTUAL util: climbs from util_start toward the ceiling (or == ceiling if start==0)
    actual = {y: (ustart + (ceiling[y] - ustart) * maturity(y)) if ustart > 0 else ceiling[y] for y in YEARS}
    used = dict(actual)                                                      # col L (no override in cache)
    realized = {y: (floored[y] * used[2025] / used[y] if split else floored[y]) for y in YEARS}
    gap = {y: realized[y] - supply[y] for y in YEARS}

    peak_gap = max(gap.values())
    peak_year = [y for y in YEARS if gap[y] == peak_gap][0]
    ov = [y for y in YEARS if gap[y] < 0]
    overshoot = ov[0] if ov else None
    bal = [y for y in YEARS if y >= peak_year and gap[y] < 30]
    balance = bal[0] if bal else None
    asym = None
    for i, y in enumerate(YEARS):
        if i and y > 2027 and realized[YEARS[i - 1]] > 0 and \
                (realized[y] - realized[YEARS[i - 1]]) / realized[YEARS[i - 1]] < 0.02:
            asym = y
            break
    return dict(tot=tot, new_eff=new_eff, fleet_eff=fleet_eff, raw=raw, floored=floored,
                supply=supply, a_share=a_share, ceiling=ceiling, actual=actual, used=used,
                realized=realized, gap=gap, peak_gap=peak_gap, peak_year=peak_year,
                overshoot=overshoot, balance=balance, asym=asym, user_scale=user_scale)


def build_xml():
    R = reference()
    rows = {}
    us = repr(R["user_scale"])

    rows[1] = [cs("A1", "EFFICIENCY OVERLAY  —  THE CORE MODEL  ·  read the flow top-to-bottom, years left-to-right", S_TITLE)]
    rows[2] = [cs("A2", "ONE model laid out as a WATERFALL: Demand → Supply → Utilization → Result, each a row, years 2025-2042 across the columns. "
                        "Pick the scenario in B4. Utilization is FORECAST, not assumed: 'util ceiling' is the achievable max; 'util ACTUAL' climbs from "
                        "~10% today (util_start) toward it by the maturity year. Realized power = demand × util(2025)/util(year) — watch it bend down as "
                        "the util row rises. Type a % into the 'util override' row to force any year.", S_ITAL)]

    rows[4] = [cs("A4", "SCENARIO   (1=Bull  ·  2=Base  ·  3=Bear  ·  4=Central)", S_LBL),
               cn("B4", PICKER_DEFAULT, S_INP_PICK),
               cf("C4", '="  active:  "&CHOOSE(B4,"Bull","Base","Bear","Central")', "  active:  Base", S_LBL)]

    # ---- HEADLINE KPI band: references the FLOW rows (gap R_GAP, realized R_REAL, supply R_SUP) ----
    gr = f"B{R_GAP}:{LASTCOL}{R_GAP}"
    yrng = f"B{YHDR}:{LASTCOL}{YHDR}"
    rng = lambda rr: f"B{rr}:{LASTCOL}{rr}"
    rows[6] = [cs("A6", "HEADLINE   ·   recomputes when you change B4", S_BANNER),
               cs("B6", "", S_BANNER), cs("C6", "", S_BANNER), cs("D6", "", S_BANNER)]
    rows[7] = [cs("A7", "PEAK GAP — demand power vs supply (GW)", S_LBL),
               cf("B7", f"=MAX({gr})", R["peak_gap"], S_KGW),
               cs("C7", "peak year ▶", S_LBL),
               cf("D7", f"=INDEX({yrng},MATCH(MAX({gr}),{gr},0))", R["peak_year"], S_KYR)]
    rows[8] = [cs("A8", "   power drawn at the peak (GW)", S_GEN),
               cf("B8", f"=INDEX({rng(R_REAL)},MATCH(MAX({gr}),{gr},0))", R["realized"][R["peak_year"]], S_KGW)]
    rows[9] = [cs("A9", "   supply at the peak (GW)", S_GEN),
               cf("B9", f"=INDEX({rng(R_SUP)},MATCH(MAX({gr}),{gr},0))", R["supply"][R["peak_year"]], S_KGW)]
    rows[10] = [cs("A10", "POWER drawn in 2042 (GW)", S_LBL),
                cf("B10", f"=ROUND({LASTCOL}{R_REAL},0)", round(R["realized"][2042]), S_KGW)]
    rows[11] = [cs("A11", "BALANCE year (gap < 30 after peak)", S_LBL),
                cf("B11", f'=IF(MIN({rng(R_BAL)})=99999,"post-2042",MIN({rng(R_BAL)}))',
                   R["balance"] if R["balance"] else "post-2042", S_KYR)]
    rows[12] = [cs("A12", "OVERSHOOT year (supply ≥ demand)", S_LBL),
                cf("B12", f'=IF(MIN({rng(R_OVS)})=99999,"post-2042",MIN({rng(R_OVS)}))',
                   R["overshoot"] if R["overshoot"] else "post-2042", S_KYR)]

    # ---- scenario inputs (levers; Bull/Base/Bear/Central in cols B-E) ----
    labels = {
        "doubling_years": "efficiency doubling (yrs)  — slower = more power",
        "fleet_lag_years": "fleet refresh lag (yrs)",
        "supply_scale": "supply build scale (x phase rates)",
        "retire_rate": "capacity retirement /yr (ITEM 9)",
        "enable_split": "credit utilization in power? (1=yes, 0=no)",
        "util_human": "util_human (interactive ceiling)",
        "util_agent": "util_agent (batch/agentic ceiling)",
        "util_start": "util_start — 2025 ACTUAL utilization",
        "util_maturity_yr": "util matures to ceiling by (year)",
        "second_wave": "ITEM 7 second-wave demand re-accel /yr (0=off)",
    }
    rows[SI_HDR] = [cs(f"A{SI_HDR}", "SCENARIO INPUTS", S_BANNER), cs(f"B{SI_HDR}", "Bull", S_COLHDR),
                    cs(f"C{SI_HDR}", "Base", S_COLHDR), cs(f"D{SI_HDR}", "Bear", S_COLHDR), cs(f"E{SI_HDR}", "Central", S_COLHDR)]
    for i, k in enumerate(SKEYS):
        r = SK0 + i
        b, c, d, e = SCEN[k]
        sty = S_YR if k == "util_maturity_yr" else S_GEN
        rows[r] = [cs(f"A{r}", labels[k], S_GEN), cn(f"B{r}", b, sty), cn(f"C{r}", c, sty),
                   cn(f"D{r}", d, sty), cn(f"E{r}", e, sty)]

    rows[SW_ROW] = [cs(f"A{SW_ROW}", "second-wave start year (ITEM 7)", S_GEN), cn(f"B{SW_ROW}", SW_START, S_YR)]
    rows[PR_HDR] = [cs(f"A{PR_HDR}", "SUPPLY PHASE RATES (per yr, editable)", S_BANNER)]
    pr_cell = {}
    for i, (cap, lbl, val) in enumerate([(2027, "2026-2027", PHASE[0]), (2030, "2028-2030", PHASE[1]),
                                         (2035, "2031-2035", PHASE[2]), (2042, "2036-2042", PHASE[3])]):
        r = PR0 + i
        rows[r] = [cs(f"A{r}", lbl, S_GEN), cn(f"B{r}", val, S_PCT)]
        pr_cell[cap] = f"$B${r}"

    # ---- active (scenario-picked) block ----
    rows[ACT_HDR] = [cs(f"A{ACT_HDR}", "ACTIVE (from scenario picker B4)", S_BANNER)]
    short = {"doubling_years": "efficiency doubling (yrs)", "fleet_lag_years": "fleet refresh lag (yrs)",
             "supply_scale": "supply build scale", "retire_rate": "capacity retirement /yr",
             "enable_split": "credit utilization?", "util_human": "util_human", "util_agent": "util_agent",
             "util_start": "util_start (2025 actual)", "util_maturity_yr": "util maturity year",
             "second_wave": "second-wave re-accel /yr"}
    act = {}
    for i, k in enumerate(SKEYS):
        r, srow = ACT0 + i, SK0 + i
        sty = S_YR if k == "util_maturity_yr" else S_GEN
        rows[r] = [cs(f"A{r}", short[k], S_GEN),
                   cf(f"B{r}", f"=CHOOSE($B$4,B{srow},C{srow},D{srow},E{srow})", SCEN[k][PICKER_DEFAULT - 1], sty)]
        act[k] = f"$B${r}"
    A_DBL, A_LAG, A_SCALE, A_RET = act["doubling_years"], act["fleet_lag_years"], act["supply_scale"], act["retire_rate"]
    A_SPLIT, A_UH, A_UA = act["enable_split"], act["util_human"], act["util_agent"]
    A_START, A_MAT, A_SW = act["util_start"], act["util_maturity_yr"], act["second_wave"]

    # ================= THE FLOW (transposed waterfall) =================
    def banner(r, text):
        rows[r] = [cs(f"A{r}", text, S_BANNER)] + [cs(f"{yc(y)}{r}", "", S_BANNER) for y in YEARS]

    def frow(r, label, vstyle, fn, cache, lstyle=S_LBL):
        cells = [cs(f"A{r}", label, lstyle)]
        for y in YEARS:
            c, pc = yc(y), (yc(y - 1) if y > 2025 else None)
            cells.append(cf(f"{c}{r}", fn(y, c, pc), cache[y], vstyle))
        rows[r] = cells

    banner(FLOW_BANNER, "THE FLOW   —   read down (the calc) and across (the years)")
    rows[YHDR] = [cs(f"A{YHDR}", "year  →", S_COLHDR)] + [cn(f"{yc(y)}{YHDR}", int(y), S_COLHDR) for y in YEARS]

    banner(R_DEMHDR, "DEMAND")
    frow(R_TOK, "tokens / day  (T)", S_GW,
         lambda y, c, pc: f"='Base'!{bcol(y)}{B_RETAIL}*{us}+'Base'!{bcol(y)}{B_ENT}+'Base'!{bcol(y)}{B_ROBO}", R["tot"])
    frow(R_NEWEFF, "  new-GPU efficiency  (idx)", S_EFF,
         lambda y, c, pc: f"=2^(({c}{YHDR}-2025)/{A_DBL})", R["new_eff"])
    frow(R_FLEET, "  fleet efficiency  (idx, vintaged)", S_EFF,
         lambda y, c, pc: "=1" if pc is None else f"=(1-1/{A_LAG})*{pc}{R_FLEET}+(1/{A_LAG})*{c}{R_NEWEFF}", R["fleet_eff"])
    frow(R_RAW, f"  ÷ eff (hw × {int(ALGO*100)}%/yr algo) = raw demand  (GW)", S_GW,
         lambda y, c, pc: f"=({c}{R_TOK}/({c}{R_FLEET}*(1+{ALGO})^({c}{YHDR}-2025)))/($B${R_TOK}/$B${R_FLEET})*{ANCHOR}", R["raw"])
    frow(R_DEM, "  demand, floored  (GW)", S_GW,
         lambda y, c, pc: (f"={c}{R_RAW}" if pc is None else
                           f"=MAX({c}{R_RAW},{pc}{R_DEM}*(1-{A_RET}),IF({c}{YHDR}>=$B${SW_ROW},"
                           f"{pc}{R_DEM}*(1+MIN(({c}{R_TOK}-{pc}{R_TOK})/{pc}{R_TOK},{A_SW})),0))"), R["floored"])

    banner(R_SUPHDR, "SUPPLY")

    def sup_fn(y, c, pc):
        if pc is None:
            return f"={ANCHOR}"
        rate = (f"IF({c}{YHDR}<=2027,{pr_cell[2027]},IF({c}{YHDR}<=2030,{pr_cell[2030]},"
                f"IF({c}{YHDR}<=2035,{pr_cell[2035]},{pr_cell[2042]})))")
        return f"={pc}{R_SUP}*(1+{rate}*{A_SCALE})"
    frow(R_SUP, "supply built  (GW)", S_GW, sup_fn, R["supply"])

    banner(R_UTILHDR, "UTILIZATION   (the new piece — watch it climb)")
    frow(R_AGENT, "  agent share of tokens", S_PCT,
         lambda y, c, pc: f"=('Base'!{bcol(y)}{B_ROBO}+'Base'!{bcol(y)}{B_ENT}*(1-1/'Base'!{bcol(y)}{B_MULT}))/{c}{R_TOK}", R["a_share"])
    frow(R_CEIL, "  util ceiling  (achievable max)", S_PCT,
         lambda y, c, pc: f"={c}{R_AGENT}*{A_UA}+(1-{c}{R_AGENT})*{A_UH}", R["ceiling"])

    def act_fn(y, c, pc):
        tx = f"MIN(1,MAX(0,({c}{YHDR}-2025)/({A_MAT}-2025)))"
        return f"=IF({A_START}=0,{c}{R_CEIL},{A_START}+({c}{R_CEIL}-{A_START})*({tx})^2*(3-2*({tx})))"
    frow(R_ACTUAL, "  util ACTUAL  (matures from util_start ↑)", S_PCT, act_fn, R["actual"])
    rows[R_OVR] = [cs(f"A{R_OVR}", "  util override  (type a % here)", S_LBL)] + [
        (cn(f"{yc(y)}{R_OVR}", OVERRIDES[y], S_INP_PCT) if y in OVERRIDES else ce(f"{yc(y)}{R_OVR}", S_INP_PCT))
        for y in YEARS]
    frow(R_USED, "  util used  (override else actual)", S_PCT,
         lambda y, c, pc: f'=IF({c}{R_OVR}="",{c}{R_ACTUAL},{c}{R_OVR})', R["used"])

    banner(R_RESHDR, "RESULT")
    frow(R_REAL, "→ REALIZED POWER  (GW)", S_GW,
         lambda y, c, pc: f'=IF(OR({A_SPLIT}=1,{c}{R_OVR}<>""),{c}{R_DEM}*$B${R_USED}/{c}{R_USED},{c}{R_DEM})', R["realized"])
    frow(R_GAP, "=  GAP  (power − supply, GW)", S_GW,
         lambda y, c, pc: f"={c}{R_REAL}-{c}{R_SUP}", R["gap"])

    # helper rows (balance / overshoot flags) — small, below the result
    frow(R_BAL, "  ·bal (helper)", S_GEN,
         lambda y, c, pc: f"=IF(AND({c}{YHDR}>=$D$7,{c}{R_GAP}<30),{c}{YHDR},99999)",
         {y: (y if (y >= R["peak_year"] and R["gap"][y] < 30) else 99999) for y in YEARS}, lstyle=S_GEN)
    frow(R_OVS, "  ·over (helper)", S_GEN,
         lambda y, c, pc: f"=IF({c}{R_GAP}<0,{c}{YHDR},99999)",
         {y: (y if R["gap"][y] < 0 else 99999) for y in YEARS}, lstyle=S_GEN)

    C0 = R_OVS + 2
    rows[C0] = [cs(f"A{C0}", "SOURCES & METHODOLOGY", S_BANNER)]
    for i, (tag, txt) in enumerate(CITES):
        r = C0 + 1 + i
        rows[r] = [cs(f"A{r}", tag, S_LBL), cs(f"B{r}", txt, S_GEN)]
    last = C0 + len(CITES)

    body = "".join(f'<row r="{n}">{"".join(rows[n])}</row>' for n in sorted(rows))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<dimension ref="A1:{LASTCOL}{last}"/>'
        '<sheetViews><sheetView workbookViewId="0">'
        # freeze the label column (A) so the line-item names stay put as you scroll across years
        f'<pane xSplit="1" topLeftCell="B1" activePane="topRight" state="frozen"/>'
        '</sheetView></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/>'
        f'<cols><col min="1" max="1" width="36"/><col min="2" max="{2 + N - 1}" width="8"/></cols>'
        '<sheetData>' + body + '</sheetData></worksheet>'
    )


def preserve_edits():
    """Read user-edited inputs from the EXISTING Efficiency Overlay sheet so a rebuild keeps
    them, instead of clobbering them with the hardcoded SCEN/PHASE/SW_START seeds. Preserves:
    the scenario-input block (incl. efficiency doubling), supply phase rates, second-wave
    start year, and the per-year util-override row. The hardcoded values are only the seed
    for a fresh workbook. Safe: read-only; falls back to seeds on any mismatch."""
    global PHASE, SW_START
    if not XLSX.exists():
        return
    try:
        import openpyxl
        ws = openpyxl.load_workbook(XLSX, data_only=True)[SHEET]
    except Exception:
        return
    num = lambda v: isinstance(v, (int, float)) and not isinstance(v, bool)
    kept = []
    # scenario inputs (rows SK0.., cols B-E = Bull / Base / Bear / Central)
    for i, k in enumerate(SKEYS):
        vals = [ws.cell(SK0 + i, c).value for c in (2, 3, 4, 5)]
        if all(num(v) for v in vals) and tuple(vals) != SCEN[k]:
            SCEN[k] = tuple(vals)
            kept.append(k)
    # supply phase rates (rows PR0.., col B)
    prs = [ws.cell(PR0 + i, 2).value for i in range(4)]
    if all(num(v) for v in prs):
        PHASE = tuple(prs)
    sw = ws.cell(SW_ROW, 2).value
    if num(sw):
        SW_START = int(sw)
    # per-year util overrides (R_OVR row, year columns)
    for y in YEARS:
        v = ws.cell(R_OVR, 2 + (y - 2025)).value
        if num(v):
            OVERRIDES[y] = float(v)
    if kept or OVERRIDES:
        print(f"preserved user edits: {', '.join(kept) or '(none)'}"
              f"{'; ' + str(len(OVERRIDES)) + ' util override(s)' if OVERRIDES else ''}")


def main():
    preserve_edits()                          # keep manual edits across rebuilds
    z = zipfile.ZipFile(XLSX)
    wb = z.read("xl/workbook.xml").decode()
    rels = z.read("xl/_rels/workbook.xml.rels").decode()
    styles = z.read("xl/styles.xml").decode()

    m = re.search(r'<sheet name="%s"[^>]*r:id="(rId\d+)"' % re.escape(SHEET), wb)
    assert m, "Efficiency Overlay sheet not found"
    rid = m.group(1)
    target = "xl/" + re.search(r'<Relationship Id="%s"[^>]*Target="([^"]+)"' % rid, rels).group(1)
    styles2, base = ensure_styles(styles)
    _set_style_globals(base)                 # resolve S_* before building the sheet
    sheet_xml = build_xml()
    minidom.parseString(sheet_xml)
    minidom.parseString(styles2)

    edited = {target: sheet_xml.encode(), "xl/styles.xml": styles2.encode()}
    tmp = str(XLSX) + ".NEW.xlsx"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for item in z.infolist():
            out.writestr(item, edited.get(item.filename, z.read(item.filename)))
    z.close()

    assert zipfile.ZipFile(tmp).testzip() is None
    import openpyxl
    wb_chk = openpyxl.load_workbook(tmp, data_only=True)
    assert SHEET in wb_chk.sheetnames

    model.EXCEL_PATH = tmp
    tok = model.build_token_demand("Base")
    cols = ["gross_tokens_T", "fleet_eff_idx", "net_compute_demand_T", "demand_gw",
            "supply_gw", "gap_gw", "cumulative_power_grid_capex_b"]
    # GOLDEN: locks model.py's COMMITTED defaults (regression guard on the engine itself).
    mv_committed = model.build_macro_gap(tok)
    h = hashlib.sha256(mv_committed[cols].round(6).to_csv().encode()).hexdigest()[:16]
    assert h == GOLDEN, f"GOLDEN MISMATCH {h}"
    # TIE-OUT: the Excel 'Base' column reproduces model.py at the SAME scenario inputs. Feed
    # model.py the Excel Base params (doubling, lag) so any owner-set Excel value that differs
    # from model.py's bare defaults still ties out. Algorithmic efficiency (10%/yr) is model.py's
    # default and matches ALGO in this builder, so it is not passed explicitly here.
    mv_base = model.build_macro_gap(tok, efficiency_doubling_years=SCEN["doubling_years"][1],
                                    fleet_lag_years=SCEN["fleet_lag_years"][1])
    R = reference()
    drift_d = max(abs(R["floored"][y] - mv_base.loc[y, "demand_gw"]) for y in YEARS)
    drift_g = max(abs(R["gap"][y] - mv_base.loc[y, "gap_gw"]) for y in YEARS)
    print(f"Base tie-out vs model.py (doubling={SCEN['doubling_years'][1]}): "
          f"max demand drift {drift_d:.4f} GW, max gap drift {drift_g:.4f} GW")
    assert drift_d < 0.5 and drift_g < 0.5, "Base column does not tie out to model.py"

    os.replace(tmp, XLSX)
    model.EXCEL_PATH = XLSX
    print(f"OK: '{SHEET}' rebuilt (Base efficiency doubling = {SCEN['doubling_years'][1]} yr); model.py golden hash {h} intact")


if __name__ == "__main__":
    main()
